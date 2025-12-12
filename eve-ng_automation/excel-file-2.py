import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
import os
from pathlib import Path

class SiteFDRGenerator:
    def __init__(self, template_path, config_path, output_dir="output_sites"):
        self.template_path = template_path
        self.config_path = config_path
        self.output_dir = output_dir
        
        # Create output directory if it doesn't exist
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)

    def load_site_data(self, sheet_name="SITE_DATA"):
        """
        Load site-specific data from velos sites_info sheet.
        Expected columns in the velos_sites_info sheet:
        - SITE_NAME: Name of the site
        - CLLI: CLLI for the hostname
        - F5-VEL-AFM-BX520 (Velos Blade): Quantity of BX520 Blades
        - F5-UPGVELQSFP28-SR4 (100 GE SFPs): Quantity of SR4 optics
        - F5-UPG-VEL-QDD-FR4 (400 GE SFPs): Quantity of FR4 optics
        - INSTALL_LOCATION_NAME: Installation location name
        - LOCATION_ADDRESS: Physical address
        - ASPN_LINK: Hyperlink to site map
        - Additional_blade: Number of additional blades (combined for both devices)
        - Existing_blade: Number of existing blades (combined for both devices)
        - router_ports: Comma-separated router ports (e.g., "4/1/c23/1, 4/1/c29/1, 5/1/c23/1")
        """
        try:
            df = pd.read_excel(self.config_path, sheet_name=sheet_name)
            print(f"Loaded {len(df)} sites from data sheet")
            return df
        except Exception as e:
            print(f"Error loading site data: {e}")
            raise

    def replace_placeholders(self, text, site_data):
        """Replace placeholders in text with site-specific data."""
        if not isinstance(text, str):
            return text
        
        # Define placeholder mappings
        replacements = {
            'CLLI+': site_data.get('CLLI', ''),
            'BLD-quantity': str(site_data.get('F5-VEL-AFM-BX520', '')),
            'SR4-quantity': str(site_data.get('F5-UPGVELQSFP28-SR4', '')),
            'FR4-quantity': str(site_data.get('F5-UPG-VEL-QDD-FR4', ''))
        }
        
        # Replace placeholders
        result = text
        for placeholder, value in replacements.items():
            result = result.replace(placeholder, value)
        
        return result

    def add_hyperlink_to_scope(self, ws_scope, site_data):
        """Add hyperlink to SCOPE OF WORK tab at A3 cell."""
        aspn_link = site_data.get('ASPN_LINK', '')
        
        if aspn_link and aspn_link != '':
            if ws_scope['A3'].value:
                current_text = ws_scope['A3'].value
                updated_text = current_text.replace('ASPN-LINK', 'Link to Site MAP')
                ws_scope['A3'] = updated_text
                ws_scope['A3'].hyperlink = aspn_link
                ws_scope['A3'].font = Font(color="0563C1", underline="single")
                print(f"  Added hyperlink to SCOPE OF WORK cell A3")

    def fill_row_black(self, ws, row_number):
        """Fill an entire row with black color for visual separation."""
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        max_col = ws.max_column
        
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_number, column=col_idx).fill = black_fill

    def add_ha_connections(self, ws_fiber, site_data):
        """Add HA interconnect connections to FIBER with X-CONNECT tab."""
        clli = site_data.get('CLLI', '')
        additional_blade = site_data.get('Additional_blade', 0)
        existing_blade = site_data.get('Existing_blade', 0)
        
        try:
            additional_blade = int(additional_blade) if pd.notna(additional_blade) else 0
            existing_blade = int(existing_blade) if pd.notna(existing_blade) else 0
        except (ValueError, TypeError):
            print(f"  Warning: Invalid blade values. Skipping HA connections.")
            return
        
        if additional_blade == 0:
            print(f"  No additional blades to add. Skipping HA connections.")
            return
        
        additional_per_device = additional_blade // 2
        existing_per_device = existing_blade // 2
        
        start_slot_index = existing_per_device
        
        firewall_01 = f"{clli}Firewall-01"
        firewall_02 = f"{clli}Firewall-02"
        
        print(f"  Adding {additional_per_device} HA connections starting from slot index {start_slot_index}")
        
        current_row = 4
        
        for i in range(additional_per_device):
            slot_number = (start_slot_index + i) * 2 + 1
            
            ws_fiber[f'D{current_row}'] = "MM MPO"
            ws_fiber[f'E{current_row}'] = "AQ"
            ws_fiber[f'I{current_row}'] = firewall_01
            ws_fiber[f'J{current_row}'] = slot_number
            ws_fiber[f'K{current_row}'] = 1
            ws_fiber[f'L{current_row}'] = "MPO"
            ws_fiber[f'BF{current_row}'] = firewall_02
            ws_fiber[f'BG{current_row}'] = slot_number
            ws_fiber[f'BH{current_row}'] = 1
            ws_fiber[f'BI{current_row}'] = "MM MPO"
            
            current_row += 1
        
        print(f"  Successfully added {additional_per_device} HA connections")
        return current_row

    # -------------------------------------------------------------------------
    # 🔥 REPLACED FUNCTION STARTS HERE — NEW add_router_connections()
    # -------------------------------------------------------------------------
    def add_router_connections(self, ws_fiber, site_data, start_row):
        """Add router interconnect connections to FIBER with X-CONNECT tab."""
        clli = site_data.get('CLLI', '')
        additional_blade = site_data.get('Additional_blade', 0)
        existing_blade = site_data.get('Existing_blade', 0)
        router_ports_str = site_data.get('router_ports', '')
        
        # Convert to integers if they're strings or floats
        try:
            additional_blade = int(additional_blade) if pd.notna(additional_blade) else 0
            existing_blade = int(existing_blade) if pd.notna(existing_blade) else 0
        except (ValueError, TypeError):
            print(f"  Warning: Invalid blade values. Skipping router connections.")
            return
        
        if additional_blade == 0:
            print(f"  No additional blades to add. Skipping router connections.")
            return
        
        # Parse router ports - handle both string and already-parsed formats
        if not router_ports_str or pd.isna(router_ports_str):
            print(f"  Warning: No router ports specified. Skipping router connections.")
            return
        
        router_ports_str = str(router_ports_str).strip()
        router_ports = [port.strip() for port in router_ports_str.split(',') if port.strip()]
        
        print(f"  Parsed {len(router_ports)} router ports: {router_ports}")
        
        if len(router_ports) < additional_blade:
            print(f"  Warning: Not enough router ports ({len(router_ports)}) for {additional_blade} connections. Using available ports.")
        
        additional_per_device = additional_blade // 2
        existing_per_device = existing_blade // 2
        
        start_slot_index = existing_per_device
        
        firewall_01 = f"{clli}Firewall-01"
        firewall_02 = f"{clli}Firewall-02"
        router_07 = f"{clli}Router-07"
        router_06 = f"{clli}Router-06"
        
        print(f"  Adding {additional_blade} router connections ({additional_per_device} per firewall)")
        
        current_row = start_row
        
        # Add black separator row before router connections
        self.fill_row_black(ws_fiber, current_row)
        current_row += 1
        
        # Firewall-01 → Router-07
        for i in range(additional_per_device):
            slot_number = (start_slot_index + i) * 2 + 1
            router_port = router_ports[i] if i < len(router_ports) else ""
            
            if router_port and '/' in router_port:
                router_slot = router_port.split('/')[0]
            else:
                router_slot = router_port
            
            print(f"    FW-01 Connection {i+1}: Slot {slot_number} -> Router Port: {router_port}, Router Slot: {router_slot}")
            
            ws_fiber[f'D{current_row}'] = "SM DUP"
            ws_fiber[f'E{current_row}'] = "YL"
            ws_fiber[f'I{current_row}'] = firewall_01
            ws_fiber[f'J{current_row}'] = slot_number
            ws_fiber[f'K{current_row}'] = 2
            ws_fiber[f'L{current_row}'] = "LC"
            ws_fiber[f'BF{current_row}'] = router_07
            ws_fiber[f'BG{current_row}'] = router_slot
            ws_fiber[f'BH{current_row}'] = router_port
            ws_fiber[f'BI{current_row}'] = "LC"
            
            current_row += 1
        
        print(f"  Added {additional_per_device} connections: {firewall_01} to {router_07}")
        
        # Separator row
        self.fill_row_black(ws_fiber, current_row)
        current_row += 1
        
        # Firewall-02 → Router-06
        for i in range(additional_per_device):
            slot_number = (start_slot_index + i) * 2 + 1
            port_index = additional_per_device + i
            router_port = router_ports[port_index] if port_index < len(router_ports) else ""
            
            if router_port and '/' in router_port:
                router_slot = router_port.split('/')[0]
            else:
                router_slot = router_port
            
            print(f"    FW-02 Connection {i+1}: Slot {slot_number} -> Router Port: {router_port}, Router Slot: {router_slot}")
            
            ws_fiber[f'D{current_row}'] = "SM DUP"
            ws_fiber[f'E{current_row}'] = "YL"
            ws_fiber[f'I{current_row}'] = firewall_02
            ws_fiber[f'J{current_row}'] = slot_number
            ws_fiber[f'K{current_row}'] = 2
            ws_fiber[f'L{current_row}'] = "LC"
            ws_fiber[f'BF{current_row}'] = router_06
            ws_fiber[f'BG{current_row}'] = router_slot
            ws_fiber[f'BH{current_row}'] = router_port
            ws_fiber[f'BI{current_row}'] = "LC"
            
            current_row += 1
        
        print(f"  Added {additional_per_device} connections: {firewall_02} to {router_06}")
        print(f"  Successfully added all router connections")
        
        return current_row
    # -------------------------------------------------------------------------
    # 🔥 END OF REPLACED FUNCTION
    # -------------------------------------------------------------------------

    def generate_site_file(self, site_data):
        site_name = site_data.get('SITE_NAME', 'UNKNOWN')
        print(f"Generating file for site: {site_name}")
        
        wb = load_workbook(self.template_path)
        
        try:
            if "SCOPE OF WORK" in wb.sheetnames:
                ws_scope = wb["SCOPE OF WORK"]
                
                if ws_scope['A3'].value:
                    original_text = ws_scope['A3'].value
                    updated_text = self.replace_placeholders(original_text, site_data)
                    ws_scope['A3'] = updated_text
                    print(f"  Updated SCOPE OF WORK cell A3")
                
                self.add_hyperlink_to_scope(ws_scope, site_data)
            
            if "KEY INFO" in wb.sheetnames:
                ws_key_info = wb["KEY INFO"]
                
                ws_key_info['B6'] = "2026-VELOS-Firewall-Blade-and-Fiber-Installation-" + str(site_data.get('SITE_NAME', ''))
                ws_key_info['B13'] = site_data.get('INSTALL_LOCATION_NAME', '')
                ws_key_info['B14'] = site_data.get('LOCATION_ADDRESS', '')
                ws_key_info['B16'] = site_data.get('CITY_STATE_ZIP', '')
                
                print(f"  Updated KEY INFO cells B6, B13, B14, B16")
            
            if "FIBER with X-CONNECT" in wb.sheetnames:
                ws_fiber = wb["FIBER with X-CONNECT"]
                
                next_row = self.add_ha_connections(ws_fiber, site_data)
                
                if next_row:
                    self.add_router_connections(ws_fiber, site_data, next_row)
            
            output_filename = f"ECI_FDR_INFORMATION_{site_name.upper()}_2026.xlsx"
            output_path = os.path.join(self.output_dir, output_filename)
            wb.save(output_path)
            
            print(f"  ✓ Successfully created: {output_filename}\n")
        
        except Exception as e:
            print(f"  ✖ Error processing site {site_name}: {e}\n")
            raise
        finally:
            wb.close()

    def generate_all_sites(self, sheet_name="SITE_DATA"):
        sites_df = self.load_site_data(sheet_name)
        
        print(f"\nStarting generation of {len(sites_df)} site files...\n")
        print("=" * 60)
        
        for idx, row in sites_df.iterrows():
            site_data = row.to_dict()
            try:
                self.generate_site_file(site_data)
            except Exception as e:
                print(f"Failed to generate file for row {idx + 1}: {e}")
                continue
        
        print("=" * 60)
        print(f"\n✓ Generation complete! Files saved in '{self.output_dir}' directory")


if __name__ == "__main__":
    generator = SiteFDRGenerator(
        template_path="ECI_FDR_INFORMATION_TEMPLATE_2026.xlsx",
        config_path="velos_sites_info.xlsx",
        output_dir="generated_sites"
    )
    
    generator.generate_all_sites(sheet_name="SITE_DATA")
