import asyncio
import openpyxl
import getpass

async def ssh_connect(ipv6, old_password, new_password):
    """
    Attempt to connect via SSH using old password, then new password.
    Returns: (result_status, password_in_use) 
    - result_status: "Login Successful" or "Login Failed"
    - password_in_use: "old password", "new password", or "Login Failed"
    """
    passwords = [old_password, new_password]
    pwd_names = ['old password', 'new password']import asyncio
import openpyxl
import getpass

async def ssh_connect(ipv6, old_password, new_password):
    """
    Attempt to connect via SSH using old password, then new password.
    Returns: (result_status, password_in_use) 
    - result_status: "Login Successful" or "Login Failed"
    - password_in_use: "old password", "new password", or "Login Failed"
    """
    passwords = [old_password, new_password]
    pwd_names = ['old password', 'new password']
    
    for pwd, pwd_name in zip(passwords, pwd_names):
        try:
            print(f"  Trying {pwd_name}...")
            
            # Create SSH connection using asyncio subprocess
            # Using ssh command with timeout and strict host key checking disabled
            process = await asyncio.create_subprocess_exec(
                'ssh',
                '-o', 'StrictHostKeyChecking=no',
                '-o', 'UserKnownHostsFile=/dev/null',
                '-o', 'ConnectTimeout=10',
                '-o', 'BatchMode=no',
                f'root@{ipv6}',
                'exit',
                stdin=asyncio.subprocess.PIPE,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            # Send password and wait for completion
            stdout, stderr = await asyncio.wait_for(
                process.communicate(input=f'{pwd}\n'.encode()),
                timeout=15
            )
            
            # Check if login was successful (exit code 0 means success)
            if process.returncode == 0:
                print(f"  ✓ Login successful with {pwd_name}")
                return ("Login Successful", pwd_name)
            else:
                print(f"  ✗ Login failed with {pwd_name}")
        
        except asyncio.TimeoutError:
            print(f"  ✗ Connection timeout with {pwd_name}")
        except Exception as e:
            print(f"  ✗ Connection error with {pwd_name}: {str(e)}")
    
    # Both passwords failed
    print(f"  ✗ Both passwords failed")
    return ("Login Failed", "Login Failed")


async def process_excel():
    """Main function to process Excel file"""
    
    # Ask for two passwords
    print("Enter old password:")
    old_password = getpass.getpass()
    print("Enter new password:")
    new_password = getpass.getpass()
    
    # Excel file path
    file_path = "C:\\path\\to\\your_file.xlsx"
    
    # Load workbook and sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Find column indices
    header_map = {}
    for cell in sheet[1]:
        if cell.value:
            header_map[cell.value.lower().strip().replace('-', ' ')] = cell.column_letter
    
    # Look for the required columns
    results_col = header_map.get('results')
    password_col = header_map.get('password in use') or header_map.get('password-in-use')
    
    if not all([results_col, password_col]):
        print("\nERROR: Required columns not found.")
        print(f"Found columns: {list(header_map.keys())}")
        print("Need: 'Results', 'password-in-use'")
        return
    
    print(f"\n✓ Using columns: Results={results_col}, Password-in-use={password_col}\n")
    
    # Process each row
    success_count = 0
    fail_count = 0
    
    for row in range(2, sheet.max_row + 1):
        ipv6 = sheet[f'B{row}'].value
        
        if not ipv6:
            continue
        
        print(f"\n{'='*60}")
        print(f"Row {row}: {ipv6}")
        print(f"{'='*60}")
        
        # Attempt SSH connection
        result_status, password_in_use = await ssh_connect(ipv6, old_password, new_password)
        
        # Write results
        sheet[f'{results_col}{row}'] = result_status
        sheet[f'{password_col}{row}'] = password_in_use
        
        if result_status == "Login Successful":
            success_count += 1
            print(f"✓ Row {row} completed successfully\n")
        else:
            fail_count += 1
            print(f"✗ Row {row} failed\n")
    
    # Save workbook
    wb.save(file_path)
    
    print("\n" + "="*60)
    print(f"✅ Completed!")
    print(f"   Successful: {success_count}")
    print(f"   Failed: {fail_count}")
    print(f"   Results saved to: {file_path}")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(process_excel())
    
    for pwd, pwd_name in zip(passwords, pwd_names):
        try:
            print(f"  Trying {pwd_name}...")
            
            # Create SSH connection using asyncio subprocess
            # Using ssh command with timeout and strict host key checking disabled
            process = await asyncio.create_subprocess_exec(
                'ssh',
                '-o', 'StrictHostKeyChecking=no',
                '-o', 'UserKnownHostsFile=/dev/null',
                '-o', 'ConnectTimeout=10',
                '-o', 'BatchMode=no',
                f'root@{ipv6}',
                'exit',
                stdin=asyncio.subprocess.PIPE,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            # Send password and wait for completion
            stdout, stderr = await asyncio.wait_for(
                process.communicate(input=f'{pwd}\n'.encode()),
                timeout=15
            )
            
            # Check if login was successful (exit code 0 means success)
            if process.returncode == 0:
                print(f"  ✓ Login successful with {pwd_name}")
                return ("Login Successful", pwd_name)
            else:
                print(f"  ✗ Login failed with {pwd_name}")
        
        except asyncio.TimeoutError:
            print(f"  ✗ Connection timeout with {pwd_name}")
        except Exception as e:
            print(f"  ✗ Connection error with {pwd_name}: {str(e)}")
    
    # Both passwords failed
    print(f"  ✗ Both passwords failed")
    return ("Login Failed", "Login Failed")


async def process_excel():
    """Main function to process Excel file"""
    
    # Ask for two passwords
    print("Enter old password:")
    old_password = getpass.getpass()
    print("Enter new password:")
    new_password = getpass.getpass()
    
    # Excel file path
    file_path = "C:\\path\\to\\your_file.xlsx"
    
    # Load workbook and sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Find column indices
    header_map = {}
    for cell in sheet[1]:
        if cell.value:
            header_map[cell.value.lower().strip().replace('-', ' ')] = cell.column_letter
    
    # Look for the required columns
    results_col = header_map.get('results')
    password_col = header_map.get('password in use') or header_map.get('password-in-use')
    
    if not all([results_col, password_col]):
        print("\nERROR: Required columns not found.")
        print(f"Found columns: {list(header_map.keys())}")
        print("Need: 'Results', 'password-in-use'")
        return
    
    print(f"\n✓ Using columns: Results={results_col}, Password-in-use={password_col}\n")
    
    # Process each row
    success_count = 0
    fail_count = 0
    
    for row in range(2, sheet.max_row + 1):
        ipv6 = sheet[f'B{row}'].value
        
        if not ipv6:
            continue
        
        print(f"\n{'='*60}")
        print(f"Row {row}: {ipv6}")
        print(f"{'='*60}")
        
        # Attempt SSH connection
        result_status, password_in_use = await ssh_connect(ipv6, old_password, new_password)
        
        # Write results
        sheet[f'{results_col}{row}'] = result_status
        sheet[f'{password_col}{row}'] = password_in_use
        
        if result_status == "Login Successful":
            success_count += 1
            print(f"✓ Row {row} completed successfully\n")
        else:
            fail_count += 1
            print(f"✗ Row {row} failed\n")
    
    # Save workbook
    wb.save(file_path)
    
    print("\n" + "="*60)
    print(f"✅ Completed!")
    print(f"   Successful: {success_count}")
    print(f"   Failed: {fail_count}")
    print(f"   Results saved to: {file_path}")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(process_excel())