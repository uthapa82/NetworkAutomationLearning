import openpyxl
import getpass
import paramiko
import socket
from openpyxl.styles import PatternFill

def ssh_connect(ip_address, username, old_password, new_password):
    """
    Attempt to connect via SSH using new password first, then old password.
    Returns: (result_status, password_in_use)
    """
    passwords = [new_password, old_password]
    pwd_names = ['new password', 'old password']
    
    for pwd, pwd_name in zip(passwords, pwd_names):
        ssh = None
        try:
            print(f"  Trying {username} with {pwd_name}...")
            
            # Create SSH client
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            # Try to connect
            ssh.connect(
                hostname=ip_address,
                username=username,
                password=pwd,
                timeout=10,
                look_for_keys=False,
                allow_agent=False
            )
            
            # If we get here, connection was successful
            print(f"  ✓ Login successful with {username}/{pwd_name}")
            ssh.close()
            return ("Login Successful", pwd_name)
            
        except paramiko.AuthenticationException:
            print(f"  ✗ Authentication failed with {username}/{pwd_name}")
            if ssh:
                ssh.close()
        except socket.timeout:
            print(f"  ✗ Connection timeout")
            if ssh:
                ssh.close()
        except Exception as e:
            print(f"  ✗ Connection error: {str(e)}")
            if ssh:
                ssh.close()
    
    # Both passwords failed for this user
    print(f"  ✗ Both passwords failed for {username}")
    return ("Login Failed", "Login Failed")


def get_passwords():
    """
    Prompt user for passwords with validation.
    Returns: (old_password, new_password)
    """
    while True:
        print("\n" + "="*60)
        print("PASSWORD ENTRY")
        print("="*60)
        
        print("\nEnter old password:")
        old_password = getpass.getpass()
        
        print("Enter new password:")
        new_password = getpass.getpass()
        
        # Validate passwords
        errors = []
        
        if not old_password or old_password.strip() == "":
            errors.append("❌ Old password cannot be blank")
        
        if not new_password or new_password.strip() == "":
            errors.append("❌ New password cannot be blank")
        
        if errors:
            print("\n" + "="*60)
            print("⚠ ERROR: Invalid password entry")
            print("="*60)
            for error in errors:
                print(error)
            print("\nPlease try again...")
            continue
        
        # Confirm passwords are different (optional warning)
        if old_password == new_password:
            print("\n⚠ WARNING: Old and new passwords are identical.")
            print("Do you want to continue? (y/n): ", end="")
            confirm = input().strip().lower()
            if confirm != 'y':
                continue
        
        print("\n✓ Passwords accepted")
        return old_password, new_password


def find_ip_column(sheet):
    """Find column containing 'IP' or 'MIP' in header"""
    for cell in sheet[1]:
        if cell.value:
            header = cell.value.strip().upper()
            if header == 'IP' or header == 'MIP':
                return cell.column_letter
    return None


def insert_columns_if_needed(sheet):
    """
    Insert required columns starting from column C if they don't exist.
    Returns dictionary mapping column purposes to column letters.
    """
    required_columns = [
        'Results',
        'user-root',
        'root-password-in-use',
        'user-admin', 
        'admin-password-in-use'
    ]
    
    # Check existing headers
    existing_headers = {}
    for cell in sheet[1]:
        if cell.value:
            existing_headers[cell.value.strip().lower()] = cell.column_letter
    
    # Determine which columns need to be inserted
    columns_to_insert = []
    column_map = {}
    
    for col_name in required_columns:
        normalized = col_name.lower().replace('-', '').replace('_', '')
        found = False
        
        # Check if column already exists (with flexible matching)
        for existing_key, existing_col in existing_headers.items():
            existing_normalized = existing_key.replace('-', '').replace('_', '').replace(' ', '')
            if normalized == existing_normalized:
                column_map[col_name] = existing_col
                found = True
                break
        
        if not found:
            columns_to_insert.append(col_name)
    
    # Insert missing columns starting at column C
    if columns_to_insert:
        print(f"\n⚠ Inserting {len(columns_to_insert)} missing column(s) at column C...")
        
        # Insert columns
        sheet.insert_cols(3, len(columns_to_insert))
        
        # Set headers for new columns
        for idx, col_name in enumerate(columns_to_insert):
            col_letter = openpyxl.utils.get_column_letter(3 + idx)
            sheet[f'{col_letter}1'] = col_name
            column_map[col_name] = col_letter
            print(f"  + Added column '{col_name}' at {col_letter}")
    
    # Fill in any existing columns that weren't inserted
    for col_name in required_columns:
        if col_name not in column_map:
            normalized = col_name.lower().replace('-', '').replace('_', '')
            for existing_key, existing_col in existing_headers.items():
                existing_normalized = existing_key.replace('-', '').replace('_', '').replace(' ', '')
                if normalized == existing_normalized:
                    column_map[col_name] = existing_col
                    break
    
    return column_map


def process_excel():
    """Main function to process Excel file"""
    
    # Get and validate passwords
    old_password, new_password = get_passwords()
    
    # Excel file path
    file_path = "C:\\path\\to\\your_file.xlsx"
    
    try:
        # Load workbook and sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {file_path}")
        print("Please update the file_path variable with the correct path.")
        return
    except Exception as e:
        print(f"\n❌ ERROR: Could not open Excel file: {str(e)}")
        return
    
    # Find IP column
    ip_col = find_ip_column(sheet)
    if not ip_col:
        print("\n❌ ERROR: Could not find column with header 'IP' or 'MIP'")
        print("\nAvailable headers:")
        for cell in sheet[1]:
            if cell.value:
                print(f"  - {cell.value}")
        return
    
    print(f"\n✓ Found IP column: {ip_col}")
    
    # Insert/find required columns
    column_map = insert_columns_if_needed(sheet)
    
    print(f"\n✓ Column mapping:")
    for name, letter in column_map.items():
        print(f"  {name}: {letter}")
    print()
    
    # Define highlight styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Process each row
    success_count = 0
    fail_count = 0
    highlight_count = 0
    
    usernames = ['root', 'admin']
    
    for row in range(2, sheet.max_row + 1):
        ip_address = sheet[f'{ip_col}{row}'].value
        
        if not ip_address:
            continue
        
        print(f"\n{'='*60}")
        print(f"Row {row}: {ip_address}")
        print(f"{'='*60}")
        
        root_success = False
        admin_success = False
        highlight_row = False
        
        # Try root user
        print("\n[ROOT USER]")
        root_result, root_pwd = ssh_connect(ip_address, 'root', old_password, new_password)
        sheet[f"{column_map['user-root']}{row}"] = 'root'
        sheet[f"{column_map['root-password-in-use']}{row}"] = root_pwd
        
        if root_result == "Login Successful":
            root_success = True
            if root_pwd == 'old password':
                highlight_row = True
                print("  ⚠ WARNING: Old password still works!")
        else:
            highlight_row = True
        
        # Try admin user
        print("\n[ADMIN USER]")
        admin_result, admin_pwd = ssh_connect(ip_address, 'admin', old_password, new_password)
        sheet[f"{column_map['user-admin']}{row}"] = 'admin'
        sheet[f"{column_map['admin-password-in-use']}{row}"] = admin_pwd
        
        if admin_result == "Login Successful":
            admin_success = True
            if admin_pwd == 'old password':
                highlight_row = True
                print("  ⚠ WARNING: Old password still works!")
        else:
            highlight_row = True
        
        # Set overall result
        if root_success or admin_success:
            sheet[f"{column_map['Results']}{row}"] = "Login Successful"
            success_count += 1
        else:
            sheet[f"{column_map['Results']}{row}"] = "Login Failed"
            fail_count += 1
        
        # Highlight row if needed
        if highlight_row:
            highlight_count += 1
            print(f"\n⚠ Highlighting row {row}")
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = yellow_fill
        
        print(f"\n{'─'*60}")
        print(f"Row {row} summary:")
        print(f"  Root: {root_result} ({root_pwd})")
        print(f"  Admin: {admin_result} ({admin_pwd})")
        print(f"  Highlighted: {'Yes' if highlight_row else 'No'}")
    
    # Save workbook
    try:
        wb.save(file_path)
        print("\n" + "="*60)
        print(f"✅ COMPLETED!")
        print(f"   Successful: {success_count}")
        print(f"   Failed: {fail_count}")
        print(f"   Highlighted: {highlight_count}")
        print(f"   Results saved to: {file_path}")
        print("="*60)
    except Exception as e:
        print(f"\n❌ ERROR: Could not save file: {str(e)}")


if __name__ == "__main__":
    try:
        process_excel()
    except KeyboardInterrupt:
        print("\n\n⚠ Script interrupted by user")
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {str(e)}")